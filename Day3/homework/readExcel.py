# coding:utf-8
__author__ = 'ljy'
#create_time = 2019.10.23

import openpyxl
import requests
import json


def readExcel():
    filename = r'E:\\code\\vip3test\\Day3\\homework\\data.xlsx'
    inwd = openpyxl.load_workbook(filename)

    sheetnames = inwd.sheetnames #获取读文件中的所有sheet，通过名字的方式
    S = []
    #将所有sheet存入一个列表
    for i in range(len(sheetnames)):
        ws = inwd[sheetnames[i]]
        S.append(ws)
    return S

def returnSheet(ws):
    #取其中一个sheet读取数据，每一行作为一个list存入L
    rows = ws.max_row
    cols = ws.max_column
    #给每一行建立一个空列表
    L = []
    for i in range(1,rows+1):
        L.append([])
    #逐行逐个取出数据
    for i in range(1,rows+1):
        for j in range(1,cols+1):
            L[i-1].append(ws.cell(i,j).value)
        if i == 10:
            break
    # for i in range(len(L)):
    #     for j in range(len(L[i])):
    #         L[i][j] = L[i][j].encode('gb18030') #将unicode转化为str
    return L

def test1(L1,L2):
    urlstr = L1[1][1]
    payload = L2[1][1]

    r = requests.post(url=urlstr,data=payload)
    if r.json()['data']['username'] == payload['username']:
        result = 'success'
        return result
    else:
        result = 'failed'
        return result

def writeResult(i,result):
    #创建一个Excel
    wb = openpyxl.Workbook()
    #创建一个result的sheet
    wb.create_sheet(index=0,title='result')
    ws = wb['result']
    #写入数据，i是行和编号，result是结果
    ws['A1'] = 'id'
    ws['B1'] = 'result'
    ws.cell(i+1,1).value = i
    ws.cell(i+1,2).value = result
    #保存
    filename = 'E:\\code\\vip3test\\Day3\\homework\\result.xlsx'
    wb.save(filename)

#调用readExcel()函数，将所有sheet存储在L中
L = readExcel()
#调用returnSheet()函数，创建sheet1和sheet2
L1 = returnSheet(L[0])
print(L1)
L2 = returnSheet(L[1])
print(L2)
#将sheet2中的数据由str转换成json，方便后面调用
for i in range(1,len(L2)):
    L2[i][1] = json.loads(L2[i][1])
#执行第一条测试用例
result = test1(L1,L2)
#将结果存入新的Excel
writeResult(1,result)

